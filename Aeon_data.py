import requests
from bs4 import BeautifulSoup
 
url = 'https://produce-web.net' #取得したいサイトのURLに変える
html = requests.get(url)
soup = BeautifulSoup(html.content, "html.parser")
 
num = []
titles = []
 
tmp_num = 1
 
for element in soup.select('.entry-title'): #取得したい要素名に変える
    num.append(tmp_num)        #配列numに1から始まる整数を格納
    titles.append(element.text)　#配列titleに取得した記事タイトルを格納
    tmp_num += 1
 
 
##以下、エクセル出力に関する部分
 
import openpyxl
 
wb = openpyxl.Workbook() #エクセルファイルを新規作成
sheet = wb.active　　　　
sheet.title = 'blog_title' #excelシート名を「blog_title」に
 
sheet["A1"].value = '新着順'
sheet["B1"].value = '記事タイトル'
 
for i in range(1, tmp_num):
    sheet.cell(column=1, row=i+1, value=num[i-1]) #配列numの要素をA列に出力
    sheet.cell(column=2, row=i+1, value=titles[i-1]) #配列titleの要素をB列に出力
 
wb.save('scraping_excel.xlsx') #「scraping_excel.xlsx」というファイル名で保存
wb.close()